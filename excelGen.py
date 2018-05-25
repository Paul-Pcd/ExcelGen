# coding=utf-8
import os
import time
import os.path
import commands
from copy import copy

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from patrol import load_conf


class ExcelFormat(object):
    """
    根据巡检脚本采集的结果生成巡检报告
    """

    def __init__(self, sourcedir):
        """
        :param sourcedir: 巡检结果目录
        """
        self.sourcedir = sourcedir
        self.source_file = None  # 巡检结果文件
        self.destination_file = None  # 要生成excel文件
        self.rb = None  # excel读
        self.wb = None  # excel写
        self.conf = load_conf()
        self.get_destination_file()
        self.patrol_text = None
        self.compute_rows = 0
        with open(self.source_file, 'r') as f:
            self.patrol_text = eval(f.read())

        if not self.conf:
            print "configure file /etc/patrol/patrol.conf not found!"
            return
        # print 'self.patrol_text', self.patrol_text
        print 'self.conf:', self.conf
        if not self.patrol_text:
            print "patrol result is null!"
            return

    def convert_status(self, status):
        return u'正常' if status == 'normal' else u'异常'

    def get_destination_file(self):
        """
        获取文件路径
        source_file: 巡检目录下最新的巡检结果文件
        destination_file: 与source_file同名的目标excel文件
        :return:
        """
        dirs = os.listdir(self.sourcedir)
        for dir_file in dirs:
            if os.path.isfile(os.path.join(self.sourcedir, dir_file)):
                if not self.source_file:
                    self.source_file = dir_file
                else:
                    if int(self.source_file[3:10]) < dir_file[3:10]:
                        self.source_file = dir_file
        filename = os.path.splitext(self.source_file)
        self.source_file = os.path.join(self.sourcedir, self.source_file)
        self.destination_file = os.path.join(self.sourcedir, 'result', filename[0] + '.xlsx')
        if not os.path.isfile(self.destination_file):
            sts, output = commands.getstatusoutput('cp %s %s' % ('template/template.xlsx', self.destination_file))
            assert sts == 0
        else:
            print '%s already exists!' % self.destination_file
            self.destination_file = os.path.join(self.sourcedir, 'result', filename[0] + '.%s.xlsx' % time.time())
            print 'new docuemnt is %s!' % self.destination_file
            sts, output = commands.getstatusoutput('cp %s %s' % ('template/template.xlsx', self.destination_file))

    def get_excel(self, source=None, destination=None):
        if source:
            self.source_file = source
            with open(self.source_file, 'r') as f:
                self.patrol_text = eval(f.read())
        if destination:
            self.destination_file = destination
        self.wb = load_workbook(self.destination_file)

    def style_range(self, ws, cell_range, format_cell):
        """
        Apply styles to a range of cells as if they were a single cell.
        :param cell_range 'B2:F4'
        """

        top = Border(top=format_cell.border.top)
        left = Border(left=format_cell.border.left)
        right = Border(right=format_cell.border.right)
        bottom = Border(bottom=format_cell.border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if format_cell.alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = copy(format_cell.alignment)

        rows = ws[cell_range]
        if format_cell.font:
            first_cell.font = copy(format_cell.font)

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if format_cell.fill:
                for c in row:
                    c.fill = copy(format_cell.fill)

    def get_host_info(self):
        """
        从巡检过结果文件构建表格数据
        :return:
        """
        patrol_result = dict()
        common_conf = self.conf.get('common')
        if not common_conf:
            print "configure file error, no [common] section!"
            return

        cluster_names = common_conf.get('cluster').split(',')
        for cluster_name in cluster_names:
            cluster_conf = self.conf.get(cluster_name, {})

            patrol_result[cluster_name] = {'compute_info': [], 'storage_info': []}
            compute_host = cluster_conf.get('compute_host').split(',')
            for ch in compute_host:
                # print ch
                info = ['', cluster_name, ch]
                patrol_info = self.patrol_text.get(cluster_name + '.' + ch, {})
                if patrol_info:
                    info.append(self.convert_status(patrol_info.get('grid_res_stat')))  # 集群环境
                    info.append(self.convert_status(patrol_info.get('vote_disk_status')))  # votedisk
                    info.append(self.convert_status(patrol_info.get('asm_dg_status', {}).get('status')))  # ASM磁盘组
                    info.append(self.convert_status(patrol_info.get('multipath_status')))  # 多路径
                    info.append(self.convert_status(patrol_info.get('ntpstat')))  # 时钟同步
                    info.append(self.convert_status(patrol_info.get('file_sys_used')))  # 文件系统
                    info.append(self.convert_status(patrol_info.get('mem_avai')))  # 内存
                    info.append(self.convert_status(patrol_info.get('ibcheck')))  # IB网络
                    info.append('')  # 物理磁盘状态
                    info.append(self.convert_status(patrol_info.get('raid_status')))  # raid组状态
                    info.append(self.convert_status(patrol_info.get('asmdisk')))  # asm磁盘
                    patrol_result[cluster_name]['compute_info'].append(info)

            storage_host = cluster_conf.get('storage_host').split(',')
            for ch in storage_host:
                # print ch
                info = ['', cluster_name, ch]
                patrol_info = self.patrol_text.get(cluster_name + '.' + ch, {})
                if patrol_info:
                    info.append(self.convert_status(patrol_info.get('ise')))  # ISE服务
                flash_status = 'normal'
                nvmemgr_status = patrol_info.get('nvmemgr_status')
                if nvmemgr_status:
                    for r in nvmemgr_status.values():
                        if r != 'normal':
                            flash_status = 'abnormal'
                            break
                info.append(self.convert_status(flash_status))  # flash
                info.append(self.convert_status(patrol_info.get('iscsi_ttx')))  # ISCSI_TTX
                info.append(self.convert_status(patrol_info.get('iscsi_np')))  # ISCSI_NP
                info.append(self.convert_status(patrol_info.get('ntpstat')))  # 时钟同步
                info.append(self.convert_status(patrol_info.get('file_sys_used')))  # 文件系统
                info.append(self.convert_status(patrol_info.get('mem_avai')))  # 内存
                info.append(self.convert_status(patrol_info.get('ibcheck')))  # IB网络
                info.append(self.convert_status(patrol_info.get('disk_hotspare_status')))  # 物理磁盘状态
                info.append(self.convert_status(patrol_info.get('raid_status')))  # raid组状态
                info.append(self.convert_status(patrol_info.get('target')))  # target服务
                patrol_result[cluster_name]['storage_info'].append(info)
        return patrol_result

    def host_sheet_format(self):
        name = u'主机'
        sheet = self.wb[name]
        if not name:
            print 'Err: template update error!'
            return

        # 获取模板格式
        compute_head = list(list(sheet.values)[0])
        storage_head = list(list(sheet.values)[2])
        head_row = sheet.row_dimensions[1]
        head_height = head_row.height
        head_cell = sheet.cell(3, 3)
        data_cell = sheet.cell(2, 3)
        data_row = sheet.row_dimensions[2]
        data_height = data_row.height

        #  删除模板行
        sheet.delete_rows(1, 3)

        #  构建数据列表
        patrol_result = self.get_host_info()
        rows = list()
        merge_len = []
        for cluster_name, cluster_info in patrol_result.items():
            compute_info = cluster_info.get('compute_info', [])
            storage_info = cluster_info.get('storage_info', [])
            cluster_row = [
                [],
                [],
                compute_head,
            ]
            cluster_row.extend(compute_info)
            storage_head[1] = cluster_name
            cluster_row.append(storage_head)
            cluster_row.extend(storage_info)
            rows.extend(cluster_row)
            merge_len.append(len(compute_info) + len(storage_info))

        merge_info = []  # 记录 merge 开始行
        for i, r in enumerate(rows):
            sheet.append(r)
            if r and r[1] == u'集群':
                merge_info.append(i + 2 + 3)

        for i, row in enumerate(sheet.iter_rows()):

            #  行头格式
            if row[2].value in (u'计算节点', u'存储节点'):
                row_dimension = sheet.row_dimensions[i + 1]
                # 行高
                row_dimension.height = head_height
                for cell in row[1:]:
                    self.cell_style_copy(cell, head_cell)
            else:
                if not row[1].value:
                    continue

                row_dimension = sheet.row_dimensions[i + 1]
                row_dimension.height = data_height

                cell = row[1]
                self.cell_style_copy(cell, head_cell)
                for cell in row[2:]:
                    self.cell_style_copy(cell, data_cell)

        # 合并单元格
        for merge_start, row_num in zip(merge_info, merge_len):
            end_row = merge_start + row_num
            cell_range = 'B' + str(merge_start) + ':' + 'B' + str(end_row)
            self.style_range(sheet, cell_range=cell_range, format_cell=head_cell)
        self.wb.save(self.destination_file)

    def get_database_info(self):
        info = {}
        # patrol_staff = u'王健'
        patrol_staff = self.conf.get('common', {}).get('patrol_staff', u'王健')
        cluster_names = self.conf.get('common', {}).get('cluster', '').split(',')
        for cluster_name in cluster_names:
            cluster_conf = self.conf.get(cluster_name, {})
            hosts = cluster_conf.get('compute_host', '').split(',')
            info[cluster_name] = {}
            info[cluster_name]['grid_res_stat'] = 'normal'
            info[cluster_name]['usable_file_mb'] = 'normal'
            info[cluster_name]['rebal'] = 'normal'
            info[cluster_name]['offline_disk'] = 'normal'
            info[cluster_name]['tablespace'] = []
            info[cluster_name]['redolog'] = []
            info[cluster_name]['awr'] = []
            info[cluster_name]['count_session'] = []
            info[cluster_name]['active_session'] = []
            for host in hosts:
                patrol_info = self.patrol_text.get(cluster_name + '.' + host, {})
                # 集群资源
                if patrol_info.get('grid_res_stat') != 'normal':
                    info[cluster_name]['grid_res_stat'] = 'abnormal'
                # DG使用率
                if patrol_info.get('asm_dg_status', {}).get('usable_file_mb') != 'normal':
                    info[cluster_name]['usable_file_mb'] = 'abnormal'
                # REBAL状态
                if patrol_info.get('asm_dg_status', {}).get('rebal') != 'normal':
                    info[cluster_name]['rebal'] = 'abnormal'
                # offline disk是否存在
                if patrol_info.get('asm_dg_status', {}).get('offline_disk') != 'normal':
                    info[cluster_name]['offline_disk'] = 'abnormal'
                # 表空间使用率
                dbtbs = patrol_info.get('tablespace_status', [])
                for r in dbtbs:
                    inst = r.get('inst', [])
                    for instinfo in inst:
                        info[cluster_name]['tablespace'].append(instinfo)

                # redo log 切换频率
                redolog = patrol_info.get('redolog_status', [])
                for r in redolog:
                    inst = r.get('inst', [])
                    for instinfo in inst:
                        info[cluster_name]['redolog'].append(instinfo)

                # 统计信息是否过期
                awr_status = patrol_info.get('awr_status', [])
                for r in awr_status:
                    inst = r.get('inst', [])
                    for instinfo in inst:
                        info[cluster_name]['awr'].append(instinfo)

                # 总会话数目
                count_seeeion = patrol_info.get('count_seeeion', [])
                for r in count_seeeion:
                    inst = r.get('inst', [])
                    for instinfo in inst:
                        info[cluster_name]['count_session'].append(instinfo)

                # 活动会话数目
                active_seeeion = patrol_info.get('active_seeeion', [])
                for r in active_seeeion:
                    inst = r.get('inst', [])
                    for instinfo in inst:
                        info[cluster_name]['active_session'].append(instinfo)
        result = {}
        for cluster_name, cluster_info in info.items():
            result[cluster_name] = []

            for key, value in cluster_info.items():
                if key in ('grid_res_stat', 'usable_file_mb', 'rebal', 'offline_disk'):
                    if key == 'grid_res_stat':
                        category_name = u'集群资源'
                        check_item = u'各资源状态'
                        check_result = u'数据库、监听资源状态均为正常'
                    elif key == 'usable_file_mb':
                        category_name = u'DG组'
                        check_item = u'使用率(Usable_file_MB>0)'
                        check_result = u'均大于0'
                    elif key == 'rebal':
                        category_name = u'DG组'
                        check_item = u'REBAL状态'
                        check_result = u'均为N'
                    elif key == 'offline_disk':
                        category_name = u'DG组'
                        check_item = u'offline disk是否存在'
                        check_result = u'均为0'
                    row = [u'', '', category_name, cluster_name, check_item, check_result,
                           self.convert_status(value),
                           patrol_staff, u'']

                    result[cluster_name].append(row)
                else:
                    if key == 'tablespace':
                        category_name = u'表空间'
                        check_item = u'使用率'
                        check_result = u'均小于90%'
                    elif key == 'redolog':
                        category_name = u'Redo Log'
                        check_item = u'切换频率'
                        check_result = u'均大于2分钟'
                    elif key == 'awr':
                        category_name = u'统计信息'
                        check_item = u'是否过期'
                        check_result = u'所有表统计信息未超过一天（部分动态过期不影响执行计划）会定期执行'
                    elif key == 'count_session':
                        category_name = u'总会话数'
                        check_item = u'总会话数值'
                        check_result = u'5000左右'
                    elif key == 'active_session':
                        category_name = u'活动会话数'
                        check_item = u'活动会话数值'
                        check_result = u'200左右'

                    for instinfo in value:
                        inst = instinfo.get('instname', '')
                        status = instinfo.get('status', '')
                        if category_name in (u'总会话数', u'活动会话数'):
                            row = ['', '', category_name, inst, check_item, status, '',
                                   patrol_staff, '']
                        else:
                            row = ['', '', category_name, inst, check_item, check_result, self.convert_status(status),
                                   patrol_staff, '']
                        result[cluster_name].append(row)
        #  排序
        sort_item = {
            u'各资源状态': 1,
            u'使用率(Usable_file_MB>0)': 2,
            u'REBAL状态': 3,
            u'offline disk是否存在': 4,
            u'使用率': 5,
            u'切换频率': 6,
            u'是否过期': 7,
            u'总会话数值': 8,
            u'活动会话数值': 9,
        }
        for name, value in result.items():
            value.sort(key=lambda t: sort_item.get(t[4]))
            for i, r in enumerate(value):
                r[1] = i + 1
        return result

    def database_sheet_format(self):
        name = u'数据库'
        sheet = self.wb[name]
        if not name:
            print 'Err: template update error!'
            return

        status_cell = sheet.cell(1, 7)
        # print 'status_cell:', status_cell.value
        data_cell = sheet.cell(1, 3)
        head1_cell = sheet.cell(3, 2)
        head2_cell = sheet.cell(5, 3)
        head1 = list(list(sheet.values)[2])
        head2 = list(list(sheet.values)[4])

        #  删除模板行
        sheet.delete_rows(1, 5)

        database_info = self.get_database_info()
        # print 'database_info:', database_info
        rows = [
        ]
        for cluster_name, cluster_info in database_info.items():
            cluster_row = [
                [],
                head1,
                [],
                head2
            ]
            cluster_row.extend(cluster_info)
            rows.extend(cluster_row)

        for r in rows:
            sheet.append(r)

        merge_start_set = {
            u'表空间': None,
            u'Redo Log': None,
            u'统计信息': None,
            u'总会话数': None,
            u'活动会话数': None,
        }
        merge_category = None
        for i, row in enumerate(sheet.iter_rows()):

            #  行头格式
            if row[1].value == u'序号':  # head 2
                for cell in row[1:]:
                    self.cell_style_copy(cell, head2_cell)
            elif row[1].value and isinstance(row[1].value, int):  # 表数据格式
                # 状态单元格
                cell = list(row).pop(6)
                self.cell_style_copy(cell, status_cell)
                for cell in row[1:]:
                    self.cell_style_copy(cell, data_cell)
            elif row[1].value:  # head 1
                # for cell in row[1:]:
                # self.cell_style_copy(cell, head1_cell)
                cell_range = 'B' + str(i + 1) + ':' + 'I' + str(i + 2)
                self.style_range(sheet, cell_range=cell_range, format_cell=head1_cell)
            #  合并单元格
            merge_start = None
            if row[2].value == u'DG组' and (not merge_start or i - merge_start > 3):
                merge_start = i + 1
                sheet.merge_cells(start_row=merge_start, start_column=3, end_row=merge_start + 2, end_column=3)
                sheet.merge_cells(start_row=merge_start, start_column=4, end_row=merge_start + 2, end_column=4)

            if row[2].value in merge_start_set.keys():
                if not merge_category:
                    merge_category = row[2].value
                    merge_start_set[row[2].value] = i + 1
                if merge_category != row[2].value:
                    merge_start_row = merge_start_set.get(merge_category)
                    # 分类列合并
                    category_cell_range = 'C' + str(merge_start_row) + ':' + 'C' + str(i)
                    self.style_range(sheet, cell_range=category_cell_range, format_cell=data_cell)
                    # sheet.merge_cells(start_row=merge_start_row, start_column=3, end_row=i, end_column=3)
                    # 检查项合并
                    cluster_cell_range = 'E' + str(merge_start_row) + ':' + 'E' + str(i)
                    self.style_range(sheet, cell_range=cluster_cell_range, format_cell=data_cell)
                    # sheet.merge_cells(start_row=merge_start_row, start_column=5, end_row=i, end_column=5)
                    # 检查结果合并
                    #check_cell_range = 'F' + str(merge_start_row) + ':' + 'F' + str(i)
                    #self.style_range(sheet, cell_range=check_cell_range, format_cell=data_cell)
                    # sheet.merge_cells(start_row=merge_start_row, start_column=6, end_row=i, end_column=6)
                    merge_category = row[2].value
                    merge_start_set[row[2].value] = i + 1
                    # print 'merge_category:', merge_category
            elif merge_category and not row[2].value:  # 单表末行特殊处理
                merge_start_row = merge_start_set.get(merge_category)
                # 分类列合并
                category_cell_range = 'C' + str(merge_start_row) + ':' + 'C' + str(i)
                self.style_range(sheet, cell_range=category_cell_range, format_cell=data_cell)
                # sheet.merge_cells(start_row=merge_start_row, start_column=3, end_row=i, end_column=3)
                # 集群/实例列合并
                cluster_cell_range = 'E' + str(merge_start_row) + ':' + 'E' + str(i)
                self.style_range(sheet, cell_range=cluster_cell_range, format_cell=data_cell)
                # sheet.merge_cells(start_row=merge_start_row, start_column=5, end_row=i, end_column=5)
                # 检查项合并
                #check_cell_range = 'F' + str(merge_start_row) + ':' + 'F' + str(i)
                #self.style_range(sheet, cell_range=check_cell_range, format_cell=data_cell)
                # sheet.merge_cells(start_row=merge_start_row, start_column=6, end_row=i, end_column=6)
                merge_category = None

        self.wb.save(self.destination_file)

    def get_log_info(self):
        log_info = {
            'Message': {},
            'ASM': {},
            'CRS': {},
            'Database': {},
        }
        cluster_names = self.conf.get('common', {}).get('cluster', '').split(',')
        for cluster_name in cluster_names:
            log_info['Message'][cluster_name] = []
            log_info['ASM'][cluster_name] = []
            log_info['CRS'][cluster_name] = []
            log_info['Database'][cluster_name] = []
            cluster_conf = self.conf.get(cluster_name, {})
            hosts = cluster_conf.get('compute_host', '').split(',')
            hosts.extend(cluster_conf.get('storage_host', '').split(','))
            for host in hosts:
                message_log_alert = self.patrol_text.get(cluster_name + '.' + host, {}).get('message_log_alert')
                if message_log_alert:
                    log_info['Message'][cluster_name].append(message_log_alert)
                asm_log_alert = self.patrol_text.get(cluster_name + '.' + host, {}).get('asm_log_alert')
                if asm_log_alert:
                    log_info['ASM'][cluster_name].append(asm_log_alert)
                crs_log_alert = self.patrol_text.get(cluster_name + '.' + host, {}).get('crs_log_alert')
                if crs_log_alert:
                    log_info['CRS'][cluster_name].append(crs_log_alert)
                oracle_log_alert = self.patrol_text.get(cluster_name + '.' + host, {}).get('oracle_log_alert')
                if oracle_log_alert:
                    log_info['Database'][cluster_name].append(oracle_log_alert)
        return log_info

    def log_sheet_format(self):
        """
        日志巡检页面
        :return:
        """
        name = u'日志'
        sheet = self.wb[name]
        if not name:
            print 'Err: template update error!'
            return

        # patrol_staff = u'王健'
        patrol_staff = self.conf.get('common', {}).get('patrol_staff', u'王健')
        status_cell = sheet.cell(1, 8)
        data_cell = sheet.cell(1, 7)

        log_info = self.get_log_info()
        # print 'log_info:', log_info
        seq = 1
        # 合并单元格
        category_col = 3  # 分类列号
        crs_col = 4  # 集群列号
        path_col = 6  # 路径列号

        category_start = 6
        crs_start = 6

        for category in ('Message', 'ASM', 'CRS', 'Database'):
            # Message
            category_len = 0
            for cluster_name, cluster_message in log_info.get(category, {}).items():
                cluster_len = len(cluster_message)
                category_len += cluster_len
                for row in cluster_message:
                    row.insert(0, u'')
                    row.insert(1, seq)  # 序号
                    row.insert(2, category)  # 检查是否包含错误，告警
                    row.insert(6, u'暂无')
                    row[7] = self.convert_status(row[7])
                    row.insert(8, patrol_staff)  # 巡检人员
                    row.insert(9, u'')  # 备注
                    seq += 1
                    sheet.append(row)
                    # print 'sheet.rows.len:', len(list(sheet.rows))
                    # print 'sheet.rows:', list(sheet.rows)
                    row = list(list(sheet.rows)[-1])
                    # print row[1].value
                    cell = row.pop(7)
                    self.cell_style_copy(cell, status_cell)
                    for cell in row[1:]:
                        self.cell_style_copy(cell, data_cell)
                # print 'crs_start:', crs_start
                # print 'cluster_len:', cluster_len
                # 集群列合并
                if cluster_len:
                    cell_range = 'D' + str(crs_start) + ':' + 'D' + str(crs_start + cluster_len - 1)
                    self.style_range(sheet, cell_range=cell_range, format_cell=data_cell)
                    # sheet.merge_cells(start_row=crs_start, start_column=crs_col, end_row=crs_start + cluster_len - 1,
                    #                  end_column=crs_col)
                crs_start = crs_start + cluster_len
            # print 'category_start:', category_start
            # print 'category_len:', category_len
            # 分类列合并
            cell_range = 'C' + str(category_start) + ':' + 'C' + str(category_start + category_len - 1)
            self.style_range(sheet, cell_range=cell_range, format_cell=data_cell)
            # sheet.merge_cells(start_row=category_start, start_column=category_col,
            #                  end_row=category_start + category_len - 1,
            #                  end_column=category_col)
            category_start = category_start + category_len

        # 清空模板行
        row = list(sheet.rows)[0]
        for cell in row:
            cell.value = ''
            cell.style = sheet.cell(1, 1).style
        self.wb.save(self.destination_file)

    def cell_style_copy(self, cell, data_cell):
        """
        单元格格式拷贝
        :param data_cell: 源格式目标单元格
        :param cell: 操作单元格
        :return:
        """
        cell.font = copy(data_cell.font)
        cell.border = copy(data_cell.border)
        cell.fill = copy(data_cell.fill)
        cell.number_format = copy(data_cell.number_format)
        cell.protection = copy(data_cell.protection)
        cell.alignment = copy(data_cell.alignment)


if __name__ == '__main__':
    sourcedir = '/umsp_data/patrol'
    ef = ExcelFormat(sourcedir)
    # ef.get_destination_file()
    # ef.get_excel(source='/umsp_data/patrol/res20180517.txt')
    #ef.get_excel(source = '/umsp_data/patrol/res20180517.txt')
    ef.get_excel()
    # print 'source file: {0}, destination_file: {1}'.format(ef.source_file, ef.destination_file)
    ef.host_sheet_format()
    ef.database_sheet_format()
    ef.log_sheet_format()
